from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename='sample.xlsx')
ws = wb['Sheet']

lineB = 'K'
sum_min = 0
sum_second = 0
start = 2
stop = 28

for i in range(start, stop):
    lineN = lineB + str(i)
    sum_min += int(ws[lineN].value[:2])
    sum_second += int(ws[lineN].value[3:])
    # print(ws[lineN].value)

carry = int(sum_second / 60)
second = sum_second - carry * 60
sum_min += carry

print("total time is", sum_min, ":", second)



