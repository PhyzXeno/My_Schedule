from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename='sample.xlsx')
ws = wb['Sheet']

def WriteExcelPos(column, row, value, ws):
    pos = column + row
    ws[pos] = value