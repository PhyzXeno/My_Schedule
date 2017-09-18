from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename='sample.xlsx')
ws = wb['Sheet']


def sum_the_time(start, stop, lineB):
    sum_min = 0
    sum_second = 0
    for i in range(start, stop):
        lineN = lineB + str(i)
        sum_min += int(ws[lineN].value.split(":")[0])
        sum_second += int(ws[lineN].value.split(":")[1])
        # print(ws[lineN].value)

    carry = int(sum_second / 60)
    second = sum_second - carry * 60
    sum_min += carry
    sum_time = str(sum_min) + ":" + str(second)
    print("total time is", sum_time, end='\n')
    total_pos = lineB + '2'
    ws[total_pos] = sum_time
    wb.save("sample.xlsx")

sum_the_time(3, 26, 'E')