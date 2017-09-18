import os
from openpyxl import Workbook
from os.path import join
import getVideoLength
import WriteExcel
from openpyxl import load_workbook

root_dir = "D:\WDS\分享的所有文件\\005_韦东山Linux_第1期视频_裸板u-boot内核文件系统驱动入门\\1期_裸机_uboot内核分析_初级驱动(2440版)\视频"
wb = Workbook()
ws = wb.active
column = 'B'
count = 0
wb = load_workbook(filename='sample.xlsx')
ws = wb['Sheet']

for root, dirs, files in os.walk(root_dir):
    for a in (join(root, name) for name in files):
        count += 1
        data = getVideoLength.getLenTime(a)
        data = float(data)
        min = int(data / 60)
        sec = int(data - min * 60)
        if(sec < 10):
            sec = "0" + str(sec)
        length = str(min) + ":" + str(sec)
        WriteExcel.WriteExcelPos(column, str(count), length, ws)

wb.save("sample.xlsx")


# for root, dirs, files in os.walk(dir):
#     for file in files:
#         count += 1
#         lineN = lineA + str(count)
#         ws[lineN] = file
#
#
# wb.save("sample.xlsx")