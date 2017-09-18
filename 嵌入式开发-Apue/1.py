from bs4 import BeautifulSoup
from openpyxl import Workbook

html_doc = open("（三）嵌入式开发-Apue - 尚观教育 - 尚观教育.html", 'rb')

soup = BeautifulSoup(html_doc, "html.parser")

span_time = ['class-hours', 'text-muted', 'pull-right', 'mrm']
div_name = ['item-content', 'pull-left']

count = 0
wb = Workbook()
ws = wb.active
lineB = 'B'
lineA = 'A'

for span in soup.find_all('span'):
    if span.get("class") == span_time:
        # print(span.get_text()[3:])
        count += 1
        lineN = lineB + str(count)
        ws[lineN] = span.get_text()[3:]

count = 0

for div in soup.find_all('div'):
    if div.get("class") == div_name:
        # print(div.get_text().strip())
        count += 1
        lineN = lineA + str(count)
        ws[lineN] = div.get_text().strip()

wb.save("sample.xlsx")
